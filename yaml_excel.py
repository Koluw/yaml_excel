import os
import random
import string

import openpyxl
import pyodbc
import xlrd
import yaml
from cryptography.fernet import Fernet

J_MAIN = 4
CONN_ARR = {'TCN': '', 'DRV': '', 'SRV': '', 'DBN': '', 'UID': '', 'PWD': '', }


def ReadXLS_ByPath(send_name):
	"""
	:param send_name: file_name to be read.
	:return: where_clause for next step (send/get response from SQL Server)
	"""
	global J_MAIN
	is_found = 0
	str_4_sql_server = ''
	try:
		wb = xlrd.open_workbook(send_name)
		sheet = wb.sheet_by_index(0)
		# Extracting number of rows
		# print(send_name + ' status[ rows:', sheet.nrows, 'cols: ', sheet.ncols, ']')
		
		for i in range(sheet.nrows):
			for j in range(sheet.ncols):
				try:
					cur_cell = str(sheet.cell_value(i, j))
					if (len(cur_cell) in (6, 8)) & (cur_cell.startswith('40')):
						
						is_found += 1
						if J_MAIN != j:
							J_MAIN = j
							print('JMain fixed')
						# print(sheet.cell_value(i, j), 'in column', J_MAIN)
						str_4_sql_server += cur_cell + ', '
				# else:
				#     print('there was no employees to check')
				except AttributeError:
					pass
	except xlrd.biffh.XLRDError:
		print('file you\'re trying to open is not excel format')
		return 'Error reading was found'
		pass
	#
	if is_found == 0:
		return 'File without needed data'
		print('there was no employees to check')  # just for local tests
	else:
		str_4_sql_server = '(' + str_4_sql_server[0:-2] + ')'
		# print(str_4_sql_server)
		# #### Here we calling another function to write data in the Excel
		Write_2Excel(send_name, str_4_sql_server)
		return str_4_sql_server


def CheckArr(sql_param='(400028, 408378, 406704)'):
	"""
	connecting to SQL Server, retrieving some data,
	:param sql_param: where_clause for check on SQL Server
	:return:
	"""
	# Create connection
	if sql_param.startswith('('):
		try:
			con = pyodbc.connect(Trusted_Connection=CONN_ARR['TCN'],
								 driver=CONN_ARR['DRV'],
								 server=CONN_ARR['SRV'],
								 database=CONN_ARR['DBN'],
								 uid=CONN_ARR['UID'],
								 pwd=CONN_ARR['PWD'])
			cur = con.cursor()
			db_cmd = "select * from v_Test where test_reason in " + sql_param
			res = cur.execute(db_cmd)
			# Do something with your result set, for example print out all the results:
			for r in res:
				print(r)
		except pyodbc.InterfaceError:
			print('connection string wasn\'t successful')
		except pyodbc.OperationalError:
			print('wow wow wow')
		return 1
	else:
		print('nothing deal with')
	pass


def Write_2Excel(send_name, some_response):
	"""
	simpliest staff: store Excel Workbook with data on last row in next column
	data we received as some_response.
	:param send_name: file_name to write in
	:param some_response: values needed to be saved in file_name
	:return:
	"""
	wb = openpyxl.load_workbook(send_name)
	sheet = wb.worksheets[0]
	
	sheet.cell(sheet.max_row, sheet.max_column + 1).value = some_response
	# print(some_response)
	wb.save(send_name)


def FindConf(config_name, selector_key):
	"""
	search for correct connectiong string for this current selector_key.
	all conn_strings are stored in yaml file.
	:param config_name: file where we wants to store new data
	:param selector_key: Selector pointed to new(existing) key in Yaml connection string
	:return: fills CONN_ARR for retreiving some data from DB.
	"""
	path_to_config = ''  # os.path.join(os.path.expandvars("%userprofile%"), 'Documents\\')
	selector_key = selector_key.upper()
	try:
		with open(path_to_config + config_name) as fn:
			configs = yaml.safe_load(fn)  # , Loader=yaml.FullLoader
			# print(configs)
			configs = configs['connectors']  # root for yaml
			for cons in configs:
				if cons == selector_key:
					curr_conn = configs[cons]
					try:
						f = Fernet(curr_conn['KEY'])
						for key in CONN_ARR:
							try:
								if key in ('DBN', 'UID', 'PWD'):
									CONN_ARR[key] = f.decrypt(str.encode(curr_conn[key])).decode()
								else:
									CONN_ARR[key] = curr_conn[key]
							except KeyError:
								print('incompatible info inside connector')
						# ##### If we want to re-create hashes for each key - next few rows shows how
						# finally:
						#     f = Fernet(curr_conn['KEY'])
						#     print(key, f.encrypt(str.encode(CONN_ARR[key])))
					except TypeError:
						print('incompatible key value for decode/encode')
				# print('this is connection string', CONN_ARR)
	# ######################### EXCEPTION BLOCK ######################### #
	except ValueError:
		print('invalid literal for int()')
	except FileNotFoundError:
		print('check if the file is present')
	except yaml.parser.ParserError:
		print('some errors during parsing config')
	except KeyError:
		print('file don\'t have corresponding value for DB_connection')
	except TypeError:
		print('trying to go through range as like through dictionary')
	except AttributeError:
		print('trying to check some Attribute which is absent')
	except yaml.composer.ComposerError:
		print('some composer problems')


def PutEntry(config_name, selector_key):
	"""
	If we need to add new connection string - this function is exactly suited for.
	:param config_name: file where we wants to store new data
	:param selector_key: Selector pointed to new(existing) key in Yaml connection string
	:return:
	"""
	path_to_config = ''  # os.path.join(os.path.expandvars("%userprofile%"), 'Documents\\programs\\200820\\')
	selector_key = selector_key.upper()
	try:
		with open(path_to_config + config_name) as fn:
			configs = yaml.safe_load(fn)  # , Loader=yaml.FullLoader
		# Do Insert new values or Update existing
		new_frnt_key = Fernet.generate_key()
		f = Fernet(new_frnt_key)
		for key in CONN_ARR:
			try:
				if key in ('DBN', 'UID', 'PWD'):
					CONN_ARR[key] = f.encrypt(str.encode(CONN_ARR[key])).decode()
			except KeyError:
				print('incompatible info inside connector')
		configs['connectors'][selector_key] = CONN_ARR.copy()
		configs['connectors'][selector_key]['KEY'] = new_frnt_key.decode()
		pass
	except yaml.parser.ParserError:
		print('some errors during parsing config')
	except yaml.composer.ComposerError:
		print('some composer problems')
	# ######################################
	# configs = {'TEST': ''}
	with open(config_name, 'w') as outfile:
		yaml.dump(configs, outfile, default_flow_style=False)


conf_name = 'data.yaml'
CONN_ARR['TCN'] = 'Yes'
CONN_ARR['DRV'] = '{SQL Server}'
CONN_ARR['SRV'] = 'sql_server'
CONN_ARR['DBN'] = 'TestingDB'
CONN_ARR['UID'] = 'testing_user'
CONN_ARR['PWD'] = 'testing_pass'
print(CONN_ARR)
PutEntry(conf_name, 'testing')
print(CONN_ARR)
FindConf(conf_name, 'test')
print(CONN_ARR)
FindConf(conf_name, 'testing')
print(CONN_ARR)
