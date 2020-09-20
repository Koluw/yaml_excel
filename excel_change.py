import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font  # colors, Fill,
from openpyxl.styles import Border, Side, Alignment, PatternFill


local_path = 'datasets\\'
file_name1 = 'b999.xlsx'
file_name2 = 'bTest202020.xlsx'


def mockup_server(server_string):
    """
    receiving string with where_clause for server, return
    :param server_string: where_clause string
    :return: df
    """
    return pd.DataFrame({'Number': [4055, 4032, 4099, 4055],
                         'Abbr': ['VVF', 'TTR', 'SWE', 'DOH'],
                         'Gear': [50, 0, 30, 50],
                        })
    pass


def merge_sources(l_src, r_src):
    return pd.merge(l_src, r_src, how='left', on='Number', left_on=None, right_on=None,
                    left_index=False, right_index=False, sort=True,
                    suffixes=('_x', '_y'), copy=True, indicator=False,
                    validate=None)


def pand_excel():
    """
    we should break file into two parts: header and data_block
    data_block we should merge with answer from server. header we should leave and add something if needed.
    This works only for only this project. For other entrance we should write another function.
    dData - result set, at first we take data with interesting block(df), send parameters to server, receiving
            answer from server and merge with df.
    dHeader - result of Header. after merge it should be upgraded - if we added new column by merge, it should be
              added also in dHeader.
    :return:
    dHeader:  Header of Excel (DataFrame)
    dData:    result set after server's answer (DataFrame)
    is_found: How many rows we have found by criteria. This criteria could be applied to this project only (Int)
    """
    HEADER_ROW = 0  # Position of Header ending
    COLUMN_POS = 0  # Position of Number's column

    xl_file = pd.ExcelFile(local_path + file_name1)

    df = xl_file.parse(xl_file.sheet_names[0], header=None)
    df.columns = ['col' + str(_ + 1) for _ in range(df.shape[1])]

    is_found = 0

    for i in range(df.shape[0]):
        j_pos = 0
        for j in df.columns:
            j_pos += 1
            cur_cell = str(df.loc[i, j])
            if (len(cur_cell) in (4, 6)) & (cur_cell.startswith('40')):
                HEADER_ROW = i
                is_found += 1
                COLUMN_POS = j_pos - 1
    # check if we have found some useful results
    if is_found == 0:
        print('There was no data for work')
        return 1, [], [], COLUMN_POS
    dHeader1 = df.loc[:HEADER_ROW - is_found]
    dHeader = dHeader1.dropna(axis=0, how='all', subset=None).copy()
    print(dHeader.shape)
    columns = ['col' + str(_ + 1) for _ in range(df.shape[1])]
    columns[COLUMN_POS] = 'Number'  # we need to change this name to use merge without additional parameters
    dHeader.columns = columns
    dData = df.loc[HEADER_ROW - is_found + 1:]
    dData.columns = columns
    number_list = '(' + str(dData['Number'].tolist()).strip('[]') + ')'
    # this string will be sent to server and from there we should get some result set.

    m_4 = mockup_server(number_list)  # mockup for answer from server
    dData = merge_sources(dData, m_4)  # refill dData with full results.
    dHeader['point1'] = 'Point_1'
    dHeader['point2'] = 'Value'

    return 0, dHeader, dData, COLUMN_POS


def save_2Excel(dataHeader, dataRows, numberColumn):
    sheetName = 'T1'
    writer = pd.ExcelWriter(local_path + file_name2, engine='openpyxl')  # , sheet_name=rs[0] xlsxwriter

    HeadRow = max(dataHeader.shape[0], 3)
    dataHeader.to_excel(writer, sheet_name=sheetName, index=False, header=False, startrow=0, startcol=0)
    dataRows.to_excel(writer, sheet_name=sheetName, index=False, header=False, startrow=HeadRow, startcol=0)

    ws = writer.book[sheetName]

    ws.sheet_view.rightToLeft = True

    # ##### Styles part  started ##### #
    ali = Alignment(wrapText=True, horizontal='center', vertical='center')

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    Regular = NamedStyle(
        name="Regular",
        font=Font(
            name="Calibri",
            size=11,
        ),
        border=thin_border,
        alignment=ali
    )

    Todo = NamedStyle(
        name="Todo",
        font=Font(
            name="Calibri",
            size=11,
        ),
        border=thin_border,
        alignment=ali,
        fill=PatternFill("solid", fgColor="000066CC")
    )

    Reply = NamedStyle(
        name="Reply",
        font=Font(
            name="Calibri",
            size=14,
            bold=True,
            color="00800080",
        ),
        border=thin_border,
        alignment=ali,
        fill=PatternFill("solid", fgColor="00FFFF00")
    )
    try:
        writer.book.add_named_style(Reply)
    except ValueError as e:
        print('Already has this one')
    try:
        writer.book.add_named_style(Regular)
    except ValueError as e:
        print('Already has Regular Style')
    try:
        writer.book.add_named_style(Todo)
    except ValueError as e:
        print('Already has Todo Style')
    # print(writer.book.style_names)
    # ##### Styles part finished ##### #
    # ################################ #

    mc = dataRows.shape[1] + 1
    mr = HeadRow + dataRows.shape[0] + 1  # to each set should be added 1, cause Excel bounds starts from 1
    j_todo = dataHeader.shape[1] - 2
    # print(mr, mc)
    # ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=4)

    for i in range(1, mr):
        for j in range(1, mc):
            ws.cell(i, j).style = "Todo" if j in (numberColumn + 1, j_todo) else "Regular"
        ws.cell(i, j).style = "Reply"
    #
    for j in range(1, mc):
        ws.merge_cells(start_row=1, start_column=j, end_row=HeadRow, end_column=j)

    for column_cells in ws.columns:
        length = 1 + max(len(as_str(cell.value)) for cell in column_cells)
        length = max(length // 5 if length > 19 else length * 1.02, 8)
        # print(length, column_cells[0].column)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length
    try:
        writer.save()
        writer.close()
    except PermissionError:
        print('The document is open already. Can\'t save new version.')


def as_str(value):
    return "" if value is None else str(value)


isNext, dHead, dData, numbRow = pand_excel()
if isNext == 0:
    save_2Excel(dHead, dData, numbRow)
else:
    print('There were few mistakes during execution.')

"""
excel = win32.gencache.EnsureDispatch('Excel.Application')
wb = excel.Workbooks.Open(r'file.xlsx')
ws = wb.Worksheets("Sheet1")
ws.Columns.AutoFit()
wb.Save()
excel.Application.Quit()
# worksheet.set_default_row(20)
    # for item, cost in expenses:
    #     ws.cell(row, col).value = item
    #     ws.cell(row, col + 1).value = cost
    #     row += 1
    # ws.cell(row, col).value = 'Total'
    # ws.cell(row, col + 1).value = '=SUM(' + get_column_letter(col + 1) + str(1) + ':' + get_column_letter(col + 1) + \
    #                           str(len(expenses)) + ')'
    #
    # workbook.save(fileName)
"""
